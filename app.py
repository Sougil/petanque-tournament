import os
from flask import Flask, request, send_file
from io import BytesIO
import pandas as pd
import random
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import logging

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='static')

def optimize_teams(players, team_size, num_matches):
    """Optimise les équipes pour maximiser la diversité des rencontres"""
    logger.debug(f"Début optimize_teams: {players} joueurs, équipes de {team_size}, {num_matches} matches")
    
    matches = []
    available_players = list(range(1, players + 1))
    
    for match in range(num_matches):
        match_teams = []
        match_players = available_players.copy()
        random.shuffle(match_players)
        
        for i in range(0, len(match_players), team_size * 2):
            if i + team_size * 2 <= len(match_players):
                team1 = match_players[i:i+team_size]
                team2 = match_players[i+team_size:i+team_size*2]
                match_teams.append((team1, team2))
        
        matches.append(match_teams)
    
    return matches

def create_match_sheet(wb, match_num, teams, team_size):
    """Crée une feuille pour un match"""
    sheet_name = f'Partie {match_num}'
    ws = wb.create_sheet(title=sheet_name)
    
    # Style des en-têtes
    header_style = Font(bold=True)
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # Configuration des colonnes
    columns = []
    for team_num in range(1, 3):  # Pour équipe 1 et 2
        for player_num in range(1, team_size + 1):
            columns.append(f'Équipe {team_num} Joueur {player_num}')
    
    columns.extend(['Résultat équipe 1', 'Résultat équipe 2', 'Points Équipe 1', 'Points Équipe 2'])
    
    # Création des en-têtes
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Remplissage des équipes
    current_row = 2
    for team1, team2 in teams:
        # Remplir équipe 1
        for i, player in enumerate(team1):
            cell = ws.cell(row=current_row, column=i+1, value=player)
            cell.alignment = Alignment(horizontal='center')
        
        # Remplir équipe 2
        for i, player in enumerate(team2):
            cell = ws.cell(row=current_row, column=team_size+i+1, value=player)
            cell.alignment = Alignment(horizontal='center')
        
        # Formules pour les points
        result1_col = len(columns) - 3
        result2_col = len(columns) - 2
        points1_col = len(columns) - 1
        points2_col = len(columns)
        
        # Formules de calcul des points
        points1_cell = ws.cell(row=current_row, column=points1_col)
        points2_cell = ws.cell(row=current_row, column=points2_col)
        
        result1_ref = f"{openpyxl.utils.get_column_letter(result1_col)}{current_row}"
        result2_ref = f"{openpyxl.utils.get_column_letter(result2_col)}{current_row}"
        
        points1_cell.value = f"={result1_ref}-{result2_ref}"
        points2_cell.value = f"={result2_ref}-{result1_ref}"
        points1_cell.alignment = Alignment(horizontal='center')
        points2_cell.alignment = Alignment(horizontal='center')
        
        current_row += 1

def create_results_sheet(wb, num_players, num_matches, team_size):
    """Crée la feuille de résultats globaux avec tri automatique"""
    ws = wb.create_sheet(title='Résultats')
    
    # En-têtes
    headers = ['Joueur', 'Victoires', 'Total points']
    header_style = Font(bold=True)
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # Définir la largeur standard pour toutes les colonnes
    for col_letter in [openpyxl.utils.get_column_letter(i) for i in range(1, len(headers) + 1)]:
        ws.column_dimensions[col_letter].width = 20
    
    # Création des en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Liste des joueurs
    for player in range(1, num_players + 1):
        row = player + 1
        
        # Numéro joueur
        ws.cell(row=row, column=1, value=player).alignment = Alignment(horizontal='center')
        
        # Formules pour les victoires (parties gagnées à 13 points)
        victories_formula_parts = []
        for match in range(1, num_matches + 1):
            sheet_name = f'Partie {match}'
            
            # Pour l'équipe 1
            for i in range(1, team_size + 1):
                col_letter = openpyxl.utils.get_column_letter(i)
                result_col = openpyxl.utils.get_column_letter(2 * team_size + 1)
                victories_formula_parts.append(
                    f"IF(AND(SUMIFS('{sheet_name}'!{result_col}:{result_col},'{sheet_name}'!{col_letter}:{col_letter},{player})=13),1,0)"
                )
            
            # Pour l'équipe 2
            for i in range(team_size + 1, 2 * team_size + 1):
                col_letter = openpyxl.utils.get_column_letter(i)
                result_col = openpyxl.utils.get_column_letter(2 * team_size + 2)
                victories_formula_parts.append(
                    f"IF(AND(SUMIFS('{sheet_name}'!{result_col}:{result_col},'{sheet_name}'!{col_letter}:{col_letter},{player})=13),1,0)"
                )
        
        # Formule pour les victoires
        ws.cell(row=row, column=2, value=f"={'+'.join(victories_formula_parts)}").alignment = Alignment(horizontal='center')
        
        # Formule pour le total des points
        points_formula_parts = []
        for match in range(1, num_matches + 1):
            sheet_name = f'Partie {match}'
            
            # Pour l'équipe 1
            for i in range(1, team_size + 1):
                col_letter = openpyxl.utils.get_column_letter(i)
                points_col = openpyxl.utils.get_column_letter(2 * team_size + 3)
                points_formula_parts.append(
                    f"SUMIF('{sheet_name}'!{col_letter}:{col_letter},{player},'{sheet_name}'!{points_col}:{points_col})"
                )
            
            # Pour l'équipe 2
            for i in range(team_size + 1, 2 * team_size + 1):
                col_letter = openpyxl.utils.get_column_letter(i)
                points_col = openpyxl.utils.get_column_letter(2 * team_size + 4)
                points_formula_parts.append(
                    f"SUMIF('{sheet_name}'!{col_letter}:{col_letter},{player},'{sheet_name}'!{points_col}:{points_col})"
                )
        
        ws.cell(row=row, column=3, value=f"={'+'.join(points_formula_parts)}").alignment = Alignment(horizontal='center')

    # Activer le filtre automatique pour permettre le tri
    ws.auto_filter.ref = f"A1:C{num_players+1}"

@app.route('/')
def index():
    try:
        return app.send_static_file('index.html')
    except Exception as e:
        logger.error(f"Erreur lors de l'accès à index.html: {str(e)}")
        return "Erreur lors du chargement de la page", 500

@app.route('/generate-tournament', methods=['POST'])
def generate_tournament():
    try:
        data = request.json
        team_type = int(data['teamType'])
        player_count = int(data['playerCount'])
        match_count = int(data['matchCount'])
        
        # Validation
        if player_count < 4:
            return {"error": "Le nombre de joueurs doit être au minimum de 4."}, 400
        if player_count < team_type * 2:
            return {"error": f"Pour un concours en {'doublette' if team_type == 2 else 'triplette'}, il faut au moins {team_type * 2} joueurs."}, 400
        if match_count < 1:
            return {"error": "Le nombre de parties doit être au minimum de 1."}, 400

        # Génération des équipes
        all_teams = optimize_teams(player_count, team_type, match_count)
        
        # Création du fichier Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Création des feuilles de match
        for match_num in range(1, match_count + 1):
            create_match_sheet(wb, match_num, all_teams[match_num-1], team_type)
        
        # Création de la feuille de résultats
        create_results_sheet(wb, player_count, match_count, team_type)
        
        # Sauvegarde
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'concours_petanque_{"doublette" if team_type == 2 else "triplette"}.xlsx'
        )
    except Exception as e:
        logger.error(f"Erreur dans generate_tournament: {str(e)}", exc_info=True)
        return {"error": f"Erreur dans generate_tournament: {str(e)}"}, 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
