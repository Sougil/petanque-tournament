import random
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

class PetanqueTournament:
    def __init__(self, team_type, num_players, num_matches):
        """
        Initialise le tournoi de pétanque
        
        :param team_type: 'doublette' ou 'triplette'
        :param num_players: nombre total de joueurs
        :param num_matches: nombre de parties à jouer
        """
        self.team_type = team_type
        self.num_players = num_players
        self.num_matches = num_matches
        self.team_size = 2 if team_type == 'doublette' else 3
        
        # Vérifie que le nombre de joueurs permet de faire des équipes complètes
        if num_players % (self.team_size * 2) != 0:
            raise ValueError(f"Le nombre de joueurs doit être un multiple de {self.team_size * 2}")
        
        # Génère la liste des joueurs
        self.players = list(range(1, num_players + 1))
        
        # Initialise les données du tournoi
        self.matches = []
        self.player_points = {player: 0 for player in self.players}
        self.player_13_points = {player: 0 for player in self.players}
    
    def generate_matches(self):
        """
        Génère les équipes aléatoirement pour chaque partie
        """
        # Mélange les joueurs
        random.shuffle(self.players)
        
        # Divise les joueurs en équipes
        for match in range(self.num_matches):
            match_teams = []
            # Divise les joueurs en deux équipes
            for team_num in range(2):
                team_start = team_num * (len(self.players) // 2)
                team_end = team_start + (len(self.players) // 2)
                team_players = self.players[team_start:team_end]
                
                # Sélectionne la taille d'équipe requise
                team = team_players[:self.team_size]
                match_teams.append(team)
            
            self.matches.append(match_teams)
        
        return self.matches
    
    def create_excel_file(self):
        """
        Crée un fichier Excel avec les résultats du tournoi
        """
        # Crée un nouveau classeur
        wb = openpyxl.Workbook()
        
        # Couleurs pour les équipes
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Style de bordure
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        
        # Crée un onglet pour chaque partie
        for match_num, match in enumerate(self.matches, 1):
            ws = wb.create_sheet(title=f'Partie {match_num}')
            
            # Définit la largeur des colonnes
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            # En-tête des équipes
            ws.cell(row=1, column=1, value=f'Equipe 1')
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).font = Font(bold=True)
            
            ws.cell(row=1, column=self.team_size + 2, value=f'Equipe 2')
            ws.cell(row=1, column=self.team_size + 2).fill = header_fill
            ws.cell(row=1, column=self.team_size + 2).font = Font(bold=True)
            
            # Ajoute les numéros de joueurs
            for team_num, team in enumerate(match):
                start_col = 1 if team_num == 0 else self.team_size + 2
                fill = yellow_fill if team_num == 0 else light_blue_fill
                
                for i, player in enumerate(team, 1):
                    cell = ws.cell(row=i+1, column=start_col, value=player)
                    cell.fill = fill
                    cell.border = border
            
            # Colonnes pour les scores et points
            headers = ['Terrain', 'Résultat Equipe 1', 'Points Equipe 1', 
                       'Résultat Equipe 2', 'Points Equipe 2']
            for i, header in enumerate(headers):
                cell = ws.cell(row=1, column=i+1, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
            
            # Ajoute une colonne "Terrain" avec un style différent
            for row in range(2, self.team_size * 2 + 2):
                terrain_cell = ws.cell(row=row, column=1, value='')
                terrain_cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                terrain_cell.border = border
            
            # Prépare les colonnes de saisie des scores avec des styles
            for row in range(2, self.team_size * 2 + 2):
                for col in range(2, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
        
        # Onglet de résultat global
        global_ws = wb.create_sheet(title='Résultat Global')
        
        # Prépare les en-têtes du classement global
        global_headers = ['Joueur', 'Total Points', 'Nb 13 Points', 'Classement', 'Ex-Aequo']
        for i, header in enumerate(global_headers, 1):
            cell = global_ws.cell(row=1, column=i, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
        
        # Supprime la feuille par défaut
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Sauvegarde le fichier
        filename = f'Tournoi_Petanque_{self.team_type.capitalize()}.xlsx'
        wb.save(filename)
        
        return filename

# Exemple d'utilisation
def main():
    # Exemple de configuration (à remplacer par une interface utilisateur)
    tournament = PetanqueTournament(
        team_type='doublette',  # ou 'triplette'
        num_players=12,
        num_matches=5
    )
    
    # Génère les équipes
    tournament.generate_matches()
    
    # Crée le fichier Excel
    filename = tournament.create_excel_file()
    print(f"Fichier généré : {filename}")

if __name__ == '__main__':
    main()
